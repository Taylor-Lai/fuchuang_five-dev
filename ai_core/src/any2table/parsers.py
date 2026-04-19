"""Parser implementations."""

from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

from any2table.core.models import (
    CanonicalDocument,
    CanonicalTable,
    DocumentBlock,
    FileAsset,
    LocationRef,
    TableCell,
    TableHeader,
    TableRow,
    TextSpan,
)

try:
    from docling.document_converter import DocumentConverter
except ImportError:  # pragma: no cover - optional dependency path
    DocumentConverter = None

W_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


def normalize_header_name(value: str) -> str:
    return "".join(value.split()).strip().lower()


class DoclingSourceParser:
    """Use Docling for source documents, then map to CanonicalDocument."""

    supported_exts = {"docx", "xlsx"}

    def __init__(self) -> None:
        self._converter = DocumentConverter() if DocumentConverter is not None else None

    def supports(self, file: FileAsset) -> bool:
        return self._converter is not None and file.role == "source" and file.ext in self.supported_exts

    def parse(self, file: FileAsset) -> CanonicalDocument:
        result = self._converter.convert(file.path)
        doc = result.document

        blocks: list[DocumentBlock] = []
        text_index: list[TextSpan] = []
        for idx, text_item in enumerate(getattr(doc, "texts", [])):
            text = (getattr(text_item, "text", None) or "").strip()
            if not text:
                continue
            location = LocationRef(doc_id=file.id, paragraph_index=idx)
            block = DocumentBlock(
                block_id=f"{file.id}#docling-p-{idx}",
                block_type=str(getattr(text_item, "label", "text")).lower(),
                text=text,
                location=location,
                attrs={"backend": "docling"},
            )
            blocks.append(block)
            text_index.append(
                TextSpan(
                    span_id=f"{file.id}#docling-span-{idx}",
                    text=text,
                    source_doc_id=file.id,
                    source_block_id=block.block_id,
                    location=location,
                )
            )

        tables: list[CanonicalTable] = []
        for table_idx, table_item in enumerate(getattr(doc, "tables", [])):
            dataframe = table_item.export_to_dataframe(doc)
            header_values = [str(col) for col in dataframe.columns]
            headers = [
                TableHeader(
                    header_id=f"{file.id}#docling-t-{table_idx}-h-{col_idx}",
                    name=value,
                    normalized_name=normalize_header_name(value),
                    col_index=col_idx,
                )
                for col_idx, value in enumerate(header_values)
            ]
            rows: list[TableRow] = []
            header_cells = [
                TableCell(
                    row_index=0,
                    col_index=col_idx,
                    value=value,
                    raw_value=value,
                    normalized_value=value,
                    location=LocationRef(doc_id=file.id, table_index=table_idx, row_index=0, col_index=col_idx),
                )
                for col_idx, value in enumerate(header_values)
            ]
            rows.append(TableRow(row_id=f"{file.id}#docling-t-{table_idx}-r-0", row_index=0, cells=header_cells))
            for row_offset, row_values in enumerate(dataframe.itertuples(index=False), start=1):
                cells: list[TableCell] = []
                for col_idx, value in enumerate(row_values):
                    cells.append(
                        TableCell(
                            row_index=row_offset,
                            col_index=col_idx,
                            value=value,
                            raw_value=value,
                            normalized_value=value,
                            location=LocationRef(doc_id=file.id, table_index=table_idx, row_index=row_offset, col_index=col_idx),
                        )
                    )
                rows.append(TableRow(row_id=f"{file.id}#docling-t-{table_idx}-r-{row_offset}", row_index=row_offset, cells=cells))
            tables.append(
                CanonicalTable(
                    table_id=f"{file.id}#docling-table-{table_idx}",
                    source_doc_id=file.id,
                    table_type="docling_table",
                    name=f"table_{table_idx}",
                    headers=headers,
                    rows=rows,
                    location=LocationRef(doc_id=file.id, table_index=table_idx),
                )
            )

        return CanonicalDocument(
            doc_id=file.id,
            file=file,
            doc_type=file.ext,
            metadata={
                "parser": "DoclingSourceParser",
                "backend": "docling",
                "text_count": len(blocks),
                "table_count": len(tables),
            },
            blocks=blocks,
            tables=tables,
            text_index=text_index,
        )


class TextParser:
    supported_exts = {"txt", "md"}

    def supports(self, file: FileAsset) -> bool:
        return file.ext in self.supported_exts

    def parse(self, file: FileAsset) -> CanonicalDocument:
        text = Path(file.path).read_text(encoding="utf-8")
        blocks = []
        text_index = []
        
        # 按换行符分割文本，创建多个block
        lines = text.split('\n')
        current_paragraph = []
        
        for i, line in enumerate(lines):
            line = line.strip()
            if line:
                current_paragraph.append(line)
            else:
                if current_paragraph:
                    paragraph_text = '\n'.join(current_paragraph)
                    location = LocationRef(doc_id=file.id, paragraph_index=len(blocks))
                    block = DocumentBlock(
                        block_id=f"{file.id}#block-{len(blocks)}",
                        block_type="paragraph",
                        text=paragraph_text,
                        location=location,
                    )
                    blocks.append(block)
                    span = TextSpan(
                        span_id=f"{file.id}#span-{len(text_index)}",
                        text=paragraph_text,
                        source_doc_id=file.id,
                        source_block_id=block.block_id,
                        location=location,
                    )
                    text_index.append(span)
                    current_paragraph = []
        
        # 处理最后一个段落
        if current_paragraph:
            paragraph_text = '\n'.join(current_paragraph)
            location = LocationRef(doc_id=file.id, paragraph_index=len(blocks))
            block = DocumentBlock(
                block_id=f"{file.id}#block-{len(blocks)}",
                block_type="paragraph",
                text=paragraph_text,
                location=location,
            )
            blocks.append(block)
            span = TextSpan(
                span_id=f"{file.id}#span-{len(text_index)}",
                text=paragraph_text,
                source_doc_id=file.id,
                source_block_id=block.block_id,
                location=location,
            )
            text_index.append(span)
        
        return CanonicalDocument(
            doc_id=file.id,
            file=file,
            doc_type=file.ext,
            metadata={"parser": "TextParser", "paragraph_count": len(blocks)},
            blocks=blocks,
            text_index=text_index,
        )


class DocxParser:
    supported_exts = {"docx"}

    def supports(self, file: FileAsset) -> bool:
        return file.ext in self.supported_exts

    def parse(self, file: FileAsset) -> CanonicalDocument:
        with ZipFile(file.path) as archive:
            data = archive.read("word/document.xml")
        root = ET.fromstring(data)

        paragraphs: list[str] = []
        blocks: list[DocumentBlock] = []
        text_index: list[TextSpan] = []
        for paragraph in root.findall(".//w:body/w:p", W_NS):
            texts = [node.text for node in paragraph.findall(".//w:t", W_NS) if node.text]
            text = "".join(texts).strip()
            if not text:
                continue
            para_index = len(paragraphs)
            paragraphs.append(text)
            location = LocationRef(doc_id=file.id, paragraph_index=para_index)
            block = DocumentBlock(
                block_id=f"{file.id}#p-{para_index}",
                block_type="paragraph",
                text=text,
                location=location,
            )
            blocks.append(block)
            text_index.append(
                TextSpan(
                    span_id=f"{file.id}#span-{para_index}",
                    text=text,
                    source_doc_id=file.id,
                    source_block_id=block.block_id,
                    location=location,
                )
            )

        table_nodes = root.findall(".//w:tbl", W_NS)
        table_contexts = paragraphs[-len(table_nodes):] if len(paragraphs) >= len(table_nodes) and table_nodes else paragraphs
        tables: list[CanonicalTable] = []
        for table_idx, table_node in enumerate(table_nodes):
            row_nodes = table_node.findall("./w:tr", W_NS)
            rows: list[TableRow] = []
            headers: list[TableHeader] = []
            for row_idx, row_node in enumerate(row_nodes):
                cells: list[TableCell] = []
                for col_idx, cell_node in enumerate(row_node.findall("./w:tc", W_NS)):
                    texts = [node.text for node in cell_node.findall(".//w:t", W_NS) if node.text]
                    cell_text = "".join(texts).strip()
                    location = LocationRef(doc_id=file.id, table_index=table_idx, row_index=row_idx, col_index=col_idx)
                    cells.append(
                        TableCell(
                            row_index=row_idx,
                            col_index=col_idx,
                            value=cell_text or None,
                            raw_value=cell_text,
                            normalized_value=cell_text or None,
                            location=location,
                        )
                    )
                    if row_idx == 0:
                        headers.append(
                            TableHeader(
                                header_id=f"{file.id}#t-{table_idx}-h-{col_idx}",
                                name=cell_text,
                                normalized_name=normalize_header_name(cell_text),
                                col_index=col_idx,
                            )
                        )
                rows.append(TableRow(row_id=f"{file.id}#t-{table_idx}-r-{row_idx}", row_index=row_idx, cells=cells))

            context_before = [table_contexts[table_idx]] if table_idx < len(table_contexts) else []
            tables.append(
                CanonicalTable(
                    table_id=f"{file.id}#table-{table_idx}",
                    source_doc_id=file.id,
                    table_type="docx_table",
                    name=f"table_{table_idx}",
                    headers=headers,
                    rows=rows,
                    context_before=context_before,
                    location=LocationRef(doc_id=file.id, table_index=table_idx),
                )
            )

        return CanonicalDocument(
            doc_id=file.id,
            file=file,
            doc_type=file.ext,
            metadata={"parser": "DocxParser", "paragraph_count": len(paragraphs), "table_count": len(tables)},
            blocks=blocks,
            tables=tables,
            text_index=text_index,
        )


class XlsxParser:
    supported_exts = {"xlsx"}

    def supports(self, file: FileAsset) -> bool:
        return file.ext in self.supported_exts

    def parse(self, file: FileAsset) -> CanonicalDocument:
        workbook = load_workbook(file.path, data_only=True)
        tables: list[CanonicalTable] = []
        for sheet_idx, worksheet in enumerate(workbook.worksheets):
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                continue
            header_values = ["" if value is None else str(value) for value in rows[0]]
            headers = [
                TableHeader(
                    header_id=f"{file.id}#s-{sheet_idx}-h-{col_idx}",
                    name=value,
                    normalized_name=normalize_header_name(value),
                    col_index=col_idx,
                )
                for col_idx, value in enumerate(header_values)
            ]
            table_rows: list[TableRow] = []
            for row_idx, row in enumerate(rows):
                cells = []
                for col_idx, value in enumerate(row):
                    cells.append(
                        TableCell(
                            row_index=row_idx,
                            col_index=col_idx,
                            value=value,
                            raw_value=value,
                            normalized_value=value,
                            location=LocationRef(doc_id=file.id, sheet=worksheet.title, row_index=row_idx, col_index=col_idx),
                        )
                    )
                table_rows.append(TableRow(row_id=f"{file.id}#s-{sheet_idx}-r-{row_idx}", row_index=row_idx, cells=cells))
            tables.append(
                CanonicalTable(
                    table_id=f"{file.id}#sheet-{sheet_idx}",
                    source_doc_id=file.id,
                    table_type="xlsx_sheet_table",
                    name=worksheet.title,
                    headers=headers,
                    rows=table_rows,
                    location=LocationRef(doc_id=file.id, sheet=worksheet.title),
                )
            )
        return CanonicalDocument(
            doc_id=file.id,
            file=file,
            doc_type=file.ext,
            metadata={"parser": "XlsxParser", "sheet_count": len(workbook.worksheets)},
            tables=tables,
        )

