import os
import uuid
from io import BytesIO
from typing import Dict, List

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class TableFiller:
    """Excel 表格自动填写器"""

    @staticmethod
    def fill_template(
        template_bytes: bytes,
        extracted_data: Dict,
        field_mapping: Dict[str, str] = None,
    ) -> BytesIO:
        """
        根据提取的数据填充 Excel 模板
        """
        try:
            # 1. 加载 Excel 模板
            wb = load_workbook(BytesIO(template_bytes))
            ws = wb.active

            print(f"📊 工作表：{ws.title}")
            print(f"📊 维度：{ws.dimensions}")

            # 2. 如果没有提供映射关系，自动匹配表头
            if not field_mapping:
                field_mapping = TableFiller._auto_match_fields(ws, extracted_data)

            print(f"🔗 字段映射：{field_mapping}")
            print(f"📦 提取数据：{extracted_data}")

            # 3. 获取表头（第一行）
            headers = []
            for cell in ws[1]:  # 第一行
                headers.append(str(cell.value).strip() if cell.value else "")

            print(f"📋 Excel 表头：{headers}")

            # 4. 填充数据（从第 2 行开始，只填充第一行数据）
            filled_count = 0
            row_idx = 2  # 只填充第 2 行

            for col_idx, header in enumerate(headers, 1):
                if not header:
                    continue

                col_letter = get_column_letter(col_idx)

                # 检查是否有映射
                if header in field_mapping:
                    extract_field = field_mapping[header]

                    # 检查是否有数据
                    if extract_field in extracted_data:
                        value = extracted_data[extract_field]
                        cell_ref = f"{col_letter}{row_idx}"
                        ws[cell_ref] = value
                        print(f"✍️  填写 {cell_ref} [{header}] = {value}")
                        filled_count += 1
                    else:
                        print(f"⚠️  字段 {extract_field} 无数据")
                else:
                    print(f"⚠️  表头 '{header}' 未匹配")

            print(f"✅ 共填充 {filled_count} 个单元格")

            # 5. 保存到 BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return output

        except Exception as e:
            import traceback

            print(f"❌ 填充失败：\n{traceback.format_exc()}")
            raise Exception(f"表格填充失败：{str(e)}")

    @staticmethod
    def _auto_match_fields(ws, extracted_data: Dict) -> Dict[str, str]:
        """
        自动匹配 Excel 表头与提取字段
        """
        try:
            # 获取第一行表头
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())

            mapping = {}

            print(f"📋 Excel 表头：{headers}")
            print(f"📋 提取字段：{list(extracted_data.keys())}")

            # 策略 1: 完全匹配
            for header in headers:
                if header in extracted_data:
                    mapping[header] = header
                    print(f"✅ 完全匹配：{header}")

            # 策略 2: 包含匹配
            for header in headers:
                if header in mapping:
                    continue
                for extract_field in extracted_data.keys():
                    # 检查是否包含关系
                    if header in extract_field or extract_field in header:
                        mapping[header] = extract_field
                        print(f"✅ 包含匹配：{header} -> {extract_field}")
                        break

            # 策略 3: 关键词匹配
            keyword_map = {
                "姓名": ["学生姓名", "姓名", "负责人"],
                "教师": ["指导教师", "指导老师", "导师", "教师"],
                "学号": ["学号", "学生学号", "编号", "ID"],
                "课程": ["课程名称", "课程", "实验课程", "科目"],
                "时间": ["实验时间", "时间", "日期", "签署日期"],
                "地点": ["实验地点", "地点", "地址"],
            }

            for header in headers:
                if header in mapping:
                    continue
                for keyword, fields in keyword_map.items():
                    if keyword in header:
                        for field in fields:
                            if field in extracted_data:
                                mapping[header] = field
                                print(f"✅ 关键词匹配：{header} ({keyword}) -> {field}")
                                break
                        if header in mapping:
                            break

            print(f"🎯 最终映射：{mapping}")
            return mapping

        except Exception as e:
            print(f"⚠️ 字段匹配警告：{str(e)}")
            return {}

    @staticmethod
    def create_template_from_fields(
        fields: List[str], output_path: str = None
    ) -> BytesIO:
        """
        根据字段列表自动生成 Excel 模板
        每个字段占一列
        """
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "数据填写表"

            # ✅ 修复：逐个字段写入不同列
            print(f"📋 创建模板，字段列表：{fields}")
            print(f"📋 字段类型：{type(fields)}")

            for col_idx, field in enumerate(fields, 1):  # 从第 1 列开始
                col_letter = get_column_letter(col_idx)
                ws[f"{col_letter}1"] = field  # 第一行是表头
                # 加粗表头
                ws[f"{col_letter}1"].font = ws[f"{col_letter}1"].font.copy(bold=True)
                print(f"  写入表头 {col_letter}1: {field}")

            # 预留 10 行数据行（从第 2 行到第 11 行）
            for row in range(2, 12):
                for col in range(1, len(fields) + 1):
                    col_letter = get_column_letter(col)
                    ws[f"{col_letter}{row}"] = ""

            # 调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width

            # 保存
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            print(f"✅ 模板创建成功，共 {len(fields)} 列")
            return output

        except Exception as e:
            import traceback

            print(f"❌ 创建模板失败：{traceback.format_exc()}")
            raise Exception(f"创建模板失败：{str(e)}")
