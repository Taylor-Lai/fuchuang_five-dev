from pathlib import Path
from typing import List
import openpyxl
from .base import DocumentChunk

class XlsxParser:
    """
    Excel 文档解析器。
    - 每张 Sheet 单独处理
    - 表头 + 每一数据行 拼接成一个 chunk（保留上下文）
    - 空行自动跳过
    """

    def __init__(self, min_chunk_length: int = 10):
        self.min_chunk_length = min_chunk_length

    def parse(self, file_path: str) -> List[DocumentChunk]:
        path = Path(file_path)
        wb = openpyxl.load_workbook(str(path), data_only=True)
        filename = path.name

        result = []
        chunk_index = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                continue

            # 第一行作为表头
            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]

            # 过滤掉全为空的表头列，记录有效列的索引
            valid_cols = [i for i, h in enumerate(headers) if h]
            if not valid_cols:
                continue

            for row in rows[1:]:
                # 跳过全空行
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                # 拼接：「表头: 值」格式，方便大模型理解
                pairs = []
                for i in valid_cols:
                    cell_val = row[i] if i < len(row) else None
                    val = str(cell_val).strip() if cell_val is not None else "(空)"
                    pairs.append(f"{headers[i]}: {val}")

                content = f"[Sheet: {sheet_name}]\n" + "\n".join(pairs)

                if len(content) < self.min_chunk_length:
                    continue

                result.append(DocumentChunk.create(
                    content=content,
                    metadata={
                        "source": filename,
                        "file_type": "xlsx",
                        "chunk_index": chunk_index,
                        "sheet_name": sheet_name,
                    }
                ))
                chunk_index += 1

        return result