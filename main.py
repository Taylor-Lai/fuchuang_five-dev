import os
import uuid
from datetime import datetime, timedelta
from typing import List, Optional
from urllib.parse import quote

from fastapi import (
    BackgroundTasks,
    Depends,
    FastAPI,
    File,
    Form,
    HTTPException,
    UploadFile,
    status,
)
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr
from sqlalchemy.orm import Session

from database import ExtractionRecord, User, get_db, init_db
from services.auth import AuthService, get_current_user
from services.db_service import DBService
from services.llm_extractor import LLMExtractor
from services.parser import DocumentParser
from services.table_filler import TableFiller

app = FastAPI(title="文档理解系统", version="1.0.0")

# ==================== 启动事件 ====================


@app.on_event("startup")
def startup_event():
    init_db()
    print("🚀 服务启动完成")


# ==================== 数据模型 ====================


class ExtractRequest(BaseModel):
    fields: List[str]


class FillTableRequest(BaseModel):
    fields: List[str]


# ==================== 数据模型 ====================


class ExtractRequest(BaseModel):
    fields: List[str]


class FillTableRequest(BaseModel):
    fields: List[str]


# ===== 新增：认证相关模型 =====
class UserCreate(BaseModel):
    username: str
    password: str
    email: EmailStr


class LoginRequest(BaseModel):
    email: str
    password: str


class Token(BaseModel):
    access_token: str
    token_type: str
    user_info: dict


# ===== 新增：修改个人资料模型 =====
class UserProfileUpdate(BaseModel):
    """修改个人资料请求模型"""
    nickname: Optional[str] = None
    gender: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None


# ==================== 基础接口 ====================


@app.get("/")
async def root():
    return {
        "message": "🚀 文档理解系统运行中",
        "version": "1.0.0",
        "docs": "/docs",
        "features": ["文档解析", "AI 提取", "自动填表", "历史记录"],
    }


@app.post("/api/upload")
async def upload_document(file: UploadFile = File(...)):
    """仅上传并解析（不提取）"""
    try:
        result = await DocumentParser.parse_file(file)
        return {
            "status": "success",
            "data": {
                "filename": result["filename"],
                "file_type": result["file_type"],
                "char_count": result["char_count"],
                "preview": (
                    result["content"][:500] + "..."
                    if len(result["content"]) > 500
                    else result["content"]
                ),
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"上传失败：{str(e)}")


# ==================== 核心提取接口 ====================


@app.post("/doc-extract/upload")
async def extract_info(
    file: UploadFile = File(..., description="源文档 (.docx, .txt 等)"),
    fields: str = Form(..., description="需要提取的字段，用逗号分隔"),
    db: Session = Depends(get_db),
):
    """
    【核心接口】上传文档 + 提取字段 → 返回提取结果（存入数据库）
    """
    try:
        # 1. 解析文档
        parse_result = await DocumentParser.parse_file(file)
        text_content = parse_result["content"]

        # 2. 处理字段列表（支持中英文逗号）
        fields_list = [
            f.strip() for f in fields.replace("，", ",").split(",") if f.strip()
        ]

        if not fields_list:
            raise HTTPException(400, "请至少指定一个填写字段")

        # 3. 调用大模型提取数据
        extracted_data = LLMExtractor.extract_info(text_content, fields_list)

        # 4. 保存记录到数据库（无论成功失败）
        status = "failed" if "error" in extracted_data else "success"
        record = DBService.save_extraction(
            db=db,
            filename=parse_result["filename"],
            file_type=parse_result["file_type"],
            fields_requested=fields_list,
            extracted_data=extracted_data,
            content_preview=text_content,
            status=status,
        )

        # 5. 检查是否提取失败
        if "error" in extracted_data:
            raise HTTPException(500, f"AI 提取失败：{extracted_data['error']}")

        return {
            "status": "success",
            "task_id": record.id,
            "filename": record.filename,
            "fields_requested": record.fields_requested,
            "extracted_data": record.extracted_data,
            "created_at": record.created_at.isoformat(),
            "query_url": f"/api/extractions/{record.id}",
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"提取失败：{str(e)}")


# ==================== 表格填写接口 ====================


@app.post("/table-fill/upload")
async def fill_table(
    template: UploadFile = File(..., description="Excel 模板文件 (.xlsx)"),
    document: UploadFile = File(..., description="源文档 (.docx, .txt, .md 等)"),
    db: Session = Depends(get_db),
):
    """
    【核心功能】上传 Excel 模板 + 文档 → 自动识别字段并填写表格
    
    自动从 Excel 模板的第一行表头中提取字段名称，
    然后从文档中提取对应数据并填写到表格中。
    """
    try:
        # 1. 校验文件类型
        if not template.filename.lower().endswith(".xlsx"):
            raise HTTPException(400, "模板文件必须是 .xlsx 格式")

        # 2. 解析文档
        parse_result = await DocumentParser.parse_file(document)
        text_content = parse_result["content"]

        # 3. 从 Excel 模板中自动提取表头字段
        template_bytes = await template.read()
        from openpyxl import load_workbook
        from io import BytesIO
        
        wb = load_workbook(BytesIO(template_bytes))
        ws = wb.active
        
        # 获取第一行表头
        fields_list = []
        for cell in ws[1]:  # 第一行
            if cell.value:
                field_name = str(cell.value).strip()
                if field_name:
                    fields_list.append(field_name)
        
        if not fields_list:
            raise HTTPException(400, "Excel 模板第一行没有找到表头字段")

        print(f"📋 自动识别的字段列表：{fields_list}")

        # 4. 调用大模型提取数据
        extracted_data = LLMExtractor.extract_info(text_content, fields_list)

        if "error" in extracted_data:
            raise HTTPException(500, f"AI 提取失败：{extracted_data['error']}")

        # 5. 填充表格
        filled_file = TableFiller.fill_template(
            template_bytes=template_bytes, extracted_data=extracted_data
        )

        # 6. 保存提取记录到数据库
        record = DBService.save_extraction(
            db=db,
            filename=parse_result["filename"],
            file_type=parse_result["file_type"],
            fields_requested=fields_list,
            extracted_data=extracted_data,
            content_preview=text_content,
            status="success",
        )

        # 7. 返回填充后的文件
        filename = f"filled_{template.filename}"
        encoded_filename = quote(filename)

        return StreamingResponse(
            filled_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"填表失败：{str(e)}")


@app.post("/table-fill/simple")
async def fill_table_simple(
    document: UploadFile = File(..., description="源文档"),
    fields: str = Form(..., description="需要提取的字段，用逗号分隔"),
    db: Session = Depends(get_db),
):
    """
    【简化版】不上传模板，自动生成 Excel 并填写
    """
    try:
        # 1. 解析文档
        parse_result = await DocumentParser.parse_file(document)
        text_content = parse_result["content"]

        # 2. 处理字段（支持中英文逗号）
        # ✅ 关键修复：确保正确分割
        fields_list = [
            f.strip() for f in fields.replace("，", ",").split(",") if f.strip()
        ]

        print(f"📋 原始 fields: {fields}")
        print(f"📋 处理后的字段列表：{fields_list}")
        print(f"📋 字段列表类型：{type(fields_list)}")
        print(f"📋 字段数量：{len(fields_list)}")

        if not fields_list:
            raise HTTPException(400, "请至少指定一个填写字段")

        # 3. 提取数据
        extracted_data = LLMExtractor.extract_info(text_content, fields_list)

        if "error" in extracted_data:
            raise HTTPException(500, f"AI 提取失败：{extracted_data['error']}")

        print(f"📦 提取的数据：{extracted_data}")

        # 4. 自动生成模板并填充
        print(f"📊 创建模板，字段：{fields_list}")
        template_bytes = TableFiller.create_template_from_fields(fields_list)

        filled_file = TableFiller.fill_template(
            template_bytes=template_bytes.getvalue(), extracted_data=extracted_data
        )

        # 5. 保存提取记录到数据库
        record = DBService.save_extraction(
            db=db,
            filename=parse_result["filename"],
            file_type=parse_result["file_type"],
            fields_requested=fields_list,
            extracted_data=extracted_data,
            content_preview=text_content,
            status="success",
        )

        # 6. 返回文件
        filename = f"auto_filled_{document.filename.split('.')[0]}.xlsx"
        encoded_filename = quote(filename)

        return StreamingResponse(
            filled_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            },
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback

        print(f"❌ 填表失败：\n{traceback.format_exc()}")
        raise HTTPException(500, f"填表失败：{str(e)}")


# ==================== 历史记录接口 ====================


@app.get("/doc-extract/search")
async def search_extractions(keyword: str, db: Session = Depends(get_db)):
    """搜索提取记录"""
    if not keyword or len(keyword) < 2:
        raise HTTPException(400, "关键词至少 2 个字符")

    records = DBService.search_extractions(db, keyword)

    return {
        "keyword": keyword,
        "total": len(records),
        "records": [
            {
                "id": r.id,
                "filename": r.filename,
                "status": r.status,
                "created_at": r.created_at.isoformat(),
                "preview": r.content_preview[:200] + "..." if r.content_preview else "",
            }
            for r in records
        ],
    }


@app.get("/doc-extract")
async def list_extractions(
    limit: int = 20,
    offset: int = 0,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),  # 此接口只有登录后才能访问
):
    """获取提取历史记录列表"""
    records = DBService.list_extractions(db, limit=limit, offset=offset)

    return {
        "total": len(records),
        "limit": limit,
        "offset": offset,
        "records": [
            {
                "id": r.id,
                "filename": r.filename,
                "file_type": r.file_type,
                "status": r.status,
                "created_at": r.created_at.isoformat(),
            }
            for r in records
        ],
    }


@app.get("/doc-extract/{record_id}")
async def get_extraction_record(record_id: str, db: Session = Depends(get_db)):
    """查询提取记录（从数据库）"""
    record = DBService.get_extraction(db, record_id)

    if not record:
        raise HTTPException(404, "记录不存在")

    return {
        "id": record.id,
        "filename": record.filename,
        "file_type": record.file_type,
        "fields_requested": record.fields_requested,
        "extracted_data": record.extracted_data,
        "status": record.status,
        "created_at": record.created_at.isoformat(),
        "updated_at": record.updated_at.isoformat() if record.updated_at else None,
    }


@app.delete("/doc-extract/{record_id}")
async def delete_extraction_record(record_id: str, db: Session = Depends(get_db)):
    """删除提取记录"""
    success = DBService.delete_extraction(db, record_id)

    if not success:
        raise HTTPException(404, "记录不存在")

    return {"message": "删除成功", "record_id": record_id}


# ==================== 认证接口 ====================


@app.post("/auth/register", response_model=dict)
async def register(user_data: UserCreate, db: Session = Depends(get_db)):
    """用户注册"""
    # 检查用户是否存在
    existing = (
        db.query(User)
        .filter((User.username == user_data.username) | (User.email == user_data.email))
        .first()
    )

    if existing:
        raise HTTPException(400, "用户名或邮箱已存在")

    # 创建用户
    user = User(
        id=str(uuid.uuid4())[:8],
        username=user_data.username,
        email=user_data.email,
        hashed_password=AuthService.get_password_hash(user_data.password),
    )

    db.add(user)
    db.commit()
    db.refresh(user)

    return {"message": "注册成功", "user_id": user.id, "username": user.username}


@app.post("/auth/login", response_model=Token)
async def login(
    login_data: LoginRequest, db: Session = Depends(get_db)
):
    """用户登录"""
    # 根据邮箱查找用户
    user = db.query(User).filter(User.email == login_data.email).first()

    if not user or not AuthService.verify_password(
        login_data.password, user.hashed_password
    ):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="邮箱或密码错误",
            headers={"WWW-Authenticate": "Bearer"},
        )

    # 创建 Token (使用用户ID作为sub，这样更通用)
    access_token = AuthService.create_access_token(
        data={"sub": user.id}, expires_delta=timedelta(minutes=30)
    )

    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user_info": {"id": user.id, "username": user.username, "email": user.email},
    }


@app.get("/user/profile", response_model=dict)
async def get_user_profile(current_user: User = Depends(get_current_user)):
    """获取个人信息"""
    return {
        "id": current_user.id,
        "username": current_user.username,
        "email": current_user.email,
        "nickname": current_user.nickname,
        "gender": current_user.gender,
        "phone": current_user.phone,
        "is_active": current_user.is_active,
    }


# 添加修改个人资料接口
@app.put("/user/profile", response_model=dict)
async def update_user_profile(
    profile_data: UserProfileUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """修改个人资料"""
    # 检查邮箱是否已被其他用户使用
    if profile_data.email and profile_data.email != current_user.email:
        existing_user = db.query(User).filter(User.email == profile_data.email).first()
        if existing_user:
            raise HTTPException(400, "邮箱已被其他用户使用")
    
    # 验证性别字段
    if profile_data.gender and profile_data.gender not in ["男", "女"]:
        raise HTTPException(400, "性别只能是'男'或'女'")
    
    # 验证手机号格式（简单验证）
    if profile_data.phone:
        import re
        phone_regex = r'^1[3-9]\d{9}$'
        if not re.match(phone_regex, profile_data.phone):
            raise HTTPException(400, "手机号格式不正确")
    
    # 更新用户信息
    if profile_data.nickname is not None:
        current_user.nickname = profile_data.nickname
    if profile_data.gender is not None:
        current_user.gender = profile_data.gender
    if profile_data.email is not None:
        current_user.email = profile_data.email
    if profile_data.phone is not None:
        current_user.phone = profile_data.phone
    
    db.commit()
    db.refresh(current_user)
    
    return {
        "message": "个人资料更新成功",
        "user": {
            "id": current_user.id,
            "username": current_user.username,
            "email": current_user.email,
            "nickname": current_user.nickname,
            "gender": current_user.gender,
            "phone": current_user.phone,
        }
    }



# ===== 新增：文档智能操作模块 =====
from services.nlp_command_parser import NLPCommandParser, ParsedCommand
from services.document_operator import DocumentOperator


class DocumentOperationRequest(BaseModel):
    """文档操作请求模型"""
    command: str  # 自然语言指令
    document: Optional[UploadFile] = None  # 可选的文档文件


class DocumentOperationResponse(BaseModel):
    """文档操作响应模型"""
    success: bool
    message: str
    result: Optional[dict] = None
    confidence: Optional[float] = None


@app.post("/doc-chat/upload", response_model=DocumentOperationResponse)
async def operate_document(
    command: str = Form(..., description="自然语言指令"),
    document: UploadFile = File(..., description="文档文件"),
):
    """
    文档智能操作接口
    
    通过自然语言指令对文档进行编辑、排版、格式调整等操作
    """
    try:
        # 1. 读取文档内容
        document_bytes = await document.read()
        
        # 2. 解析文档内容（用于上下文）
        # 读取文档内容作为上下文，以便LLM能够理解文档的实际内容
        try:
            from docx import Document
            from io import BytesIO
            doc = Document(BytesIO(document_bytes))
            document_content = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
        except Exception:
            document_content = ""  # 如果解析失败，使用空字符串
        
        # 3. 解析自然语言指令
        parsed_command = NLPCommandParser.parse_command(command, document_content)  # 传递文档内容作为上下文
        
        # 4. 执行文档操作
        operator = DocumentOperator(document_bytes)
        operation_result = operator.execute_command(parsed_command)
        
        # 5. 返回修改后的文档
        if operation_result["success"]:
            modified_document = operator.get_modified_document()
            
            # 保存到临时文件
            import tempfile
            import os
            
            with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as tmp_file:
                tmp_file.write(modified_document)
                tmp_file.flush()
                tmp_filename = tmp_file.name
            
            # 返回文件
            from fastapi.responses import FileResponse
            return FileResponse(
                tmp_filename,
                media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                filename=f"modified_{document.filename}"
            )
        else:
            return operation_result
            
    except Exception as e:
        raise HTTPException(500, f"文档操作失败: {str(e)}")

@app.get("/api/document/command-examples", response_model=dict)
async def get_command_examples():
    """
    获取常用指令示例
    """
    examples = NLPCommandParser.get_command_examples()
    return {
        "categories": examples,
        "total": len(examples)
    }


@app.post("/api/document/preview", response_model=dict)
async def preview_operation(
    command: str = Form(..., description="自然语言指令"),
    document: UploadFile = File(..., description="文档文件"),
):
    """
    预览操作效果
    
    不实际修改文档，只返回操作预览结果
    """
    try:
        # 1. 读取文档内容
        document_bytes = await document.read()
        
        # 2. 解析自然语言指令（不需要文档内容作为上下文）
        parsed_command = NLPCommandParser.parse_command(command, "")
        
        # 3. 执行操作并获取预览
        operator = DocumentOperator(document_bytes)
        operation_result = operator.execute_command(parsed_command)
        
        # 4. 获取文档预览
        preview_text = operator.get_document_preview()
        
        return {
            "command": command,
            "parsed_operation": {
                "type": parsed_command.operation_type.value,
                "params": parsed_command.params,
                "confidence": parsed_command.confidence
            },
            "operation_result": operation_result,
            "document_preview": preview_text
        }
        
    except Exception as e:
        raise HTTPException(500, f"预览失败: {str(e)}")